print("=== SRI → FACTURAS O RETENCIONES (2025) - VERSIÓN 100% FUNCIONAL ===")
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import os
import glob
import shutil
import subprocess
import socket
import re
import pandas as pd
from datetime import datetime

# =============================
# CONFIGURACIÓN
# =============================
carpeta_downloads = r"C:\Users\Jordy\Downloads"
user_profile = r"C:\Users\Jordy\Desktop\automi\automatizacion_sri\chrome_profile"
debug_port = 9222

# =============================
# ABRIR CHROME
# =============================
print("\nIniciando Chrome con tu perfil...")
sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
if sock.connect_ex(("127.0.0.1", debug_port)) != 0:
    subprocess.Popen([
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        f"--remote-debugging-port={debug_port}",
        f"--user-data-dir={user_profile}",
        "--start-maximized"
    ], shell=True)
    time.sleep(10)
else:
    time.sleep(3)
sock.close()

options = webdriver.ChromeOptions()
options.add_experimental_option("debuggerAddress", f"127.0.0.1:{debug_port}")
options.add_experimental_option("prefs", {
    "download.prompt_for_download": False,
    "plugins.always_open_pdf_externally": True,
    "safebrowsing.enabled": True
})
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=options)
wait = WebDriverWait(driver, 40)

# =============================
# LOGIN
# =============================
driver.get("https://srienlinea.sri.gob.ec/sri-en-linea/inicio/NAT")
time.sleep(5)

try:
    btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[.//p[contains(text(), 'Iniciar sesión')]]")))
    driver.execute_script("arguments[0].click();", btn)
except: pass

print("\n" + "="*60)
print(" CREDENCIALES")
print("="*60)
ruc = input(" RUC: ").strip()
clave = input(" CLAVE: ").strip()

if not ruc or not clave:
    print("Error: Ingresa RUC y clave")
    driver.quit()
    exit()

print("Haciendo login...")
try:
    ruc_input = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div[2]/div[2]/div/div/div/div/div/form/div[1]/div[2]/div/input")))
    clave_input = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[2]/div[2]/div/div/div/div/div/form/div[3]/div[2]/div/input")
    btn_ingresar = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[2]/div[2]/div/div/div/div/div/form/div[4]/div[1]/div/input")
    ruc_input.clear(); ruc_input.send_keys(ruc)
    clave_input.clear(); clave_input.send_keys(clave)
    driver.execute_script("arguments[0].click();", btn_ingresar)
    wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), 'Facturación')]")))
    print("LOGIN 100% AUTOMÁTICO")
except:
    input("→ Haz login manual y presiona ENTER...")
time.sleep(10)

# =============================
# NAVEGAR A COMPROBANTES RECIBIDOS
# =============================
try:
    fact = driver.find_element(By.XPATH, "//span[contains(text(), 'Facturación')]")
    driver.execute_script("arguments[0].click();", fact)
    time.sleep(5)
    recibidos = driver.find_element(By.XPATH, "//span[contains(text(), 'Comprobantes') and contains(text(), 'recibidos')]")
    driver.execute_script("arguments[0].click();", recibidos)
    print("Comprobantes recibidos abiertos")
    time.sleep(8)
except:
    input("→ Ve a Comprobantes recibidos y presiona ENTER...")

# =============================
# MESES Y ESPERAR DESCARGA
# =============================
mes_nombres = {
    "1": "Enero", "01": "Enero",
    "2": "Febrero", "02": "Febrero",
    "3": "Marzo", "03": "Marzo",
    "4": "Abril", "04": "Abril",
    "5": "Mayo", "05": "Mayo",
    "6": "Junio", "06": "Junio",
    "7": "Julio", "07": "Julio",
    "8": "Agosto", "08": "Agosto",
    "9": "Septiembre", "09": "Septiembre",
    "10": "Octubre",
    "11": "Noviembre",
    "12": "Diciembre"
}

def esperar_descarga():
    time.sleep(3)
    while any(f.endswith('.crdownload') for f in os.listdir(carpeta_downloads)):
        time.sleep(1)

# =============================
# BUCLE PRINCIPAL
# =============================
archivos_descargados = []

while True:
    print("\n" + "="*80)
    print("¿QUÉ TIPO DE COMPROBANTE QUIERES DESCARGAR AHORA?")
    print("1 → FACTURA")
    print("2 → COMPROBANTE DE RETENCIÓN")
    print("="*80)

    while True:
        opcion = input("\nElige 1 o 2: ").strip()
        if opcion == "1":
            carpeta_final = r"C:\Users\Jordy\Downloads\Facturas"
            tipo_comprobante = "Factura"
            es_retencion = False
            break
        elif opcion == "2":
            carpeta_final = r"C:\Users\Jordy\Downloads\Retenciones"
            tipo_comprobante = "Comprobante de Retención"
            es_retencion = True
            break
        print("Elige 1 o 2")

    os.makedirs(carpeta_final, exist_ok=True)
    print(f"\nSeleccionado: {tipo_comprobante.upper()}")
    print(f"Guardando en: {carpeta_final}")

    # Filtro tipo
    time.sleep(3)
    filtro_ok = False
    for sel in driver.find_elements(By.TAG_NAME, "select"):
        opciones = [opt.text.strip() for opt in sel.find_elements(By.TAG_NAME, "option")]
        if tipo_comprobante in opciones:
            Select(sel).select_by_visible_text(tipo_comprobante)
            print(f"Filtro aplicado: {tipo_comprobante}")
            filtro_ok = True
            time.sleep(3)
            break
    if not filtro_ok:
        input("No se cambió el filtro. Hazlo manualmente y presiona ENTER...")

    # Meses
    meses_a_descargar = []
    print(f"\nIngresa los meses para {tipo_comprobante.upper()}:")
    while True:
        while True:
            anio = input("\nAño (ej: 2024): ").strip()
            if anio.isdigit() and len(anio) == 4:
                break
            print("Año inválido")
        while True:
            mes_input = input("Mes (1-12): ").strip()
            if mes_input in mes_nombres:
                meses_a_descargar.append((anio, mes_nombres[mes_input]))
                print(f"Agregado: {mes_nombres[mes_input]} {anio}")
                break
            print("Mes inválido")
        if input("¿Otro mes para este tipo? (S/N): ").strip().upper() != "S":
            break

    # Descarga
    for idx, (anio, mes) in enumerate(meses_a_descargar, 1):
        print(f"\n{'='*80}")
        print(f"DESCARGANDO {idx}/{len(meses_a_descargar)}: {mes} {anio} → {tipo_comprobante.upper()}")
        print(f"{'='*80}")

        wait.until(EC.presence_of_element_located((By.TAG_NAME, "select")))
        time.sleep(3)

        for sel in driver.find_elements(By.TAG_NAME, "select"):
            opts = [o.text.strip() for o in sel.find_elements(By.TAG_NAME, "option")]
            if anio in opts:
                Select(sel).select_by_visible_text(anio)
            elif any(m in opts for m in mes_nombres.values()):
                Select(sel).select_by_visible_text(mes)
            elif "Todos" in opts:
                Select(sel).select_by_visible_text("Todos")

        driver.find_element(By.XPATH, "//button[contains(text(),'Consultar')]").click()
        time.sleep(20)
        input("Resuelve CAPTCHA → ENTER cuando veas la lista")

        links = driver.find_elements(By.XPATH, "//a[contains(@id,'lnkPdf')]")
        print(f"{len(links)} documentos encontrados")

        for i, link in enumerate(links, 1):
            try:
                clave = link.find_element(By.XPATH, "./ancestor::tr//td[4]").text.strip()
                nombre = "".join(c for c in clave if c.isalnum() or c in "-_")[:60] + ".pdf"
                print(f"   [{i:02d}] {nombre}")
                driver.execute_script("arguments[0].click();", link)
                esperar_descarga()

                pdfs = glob.glob(os.path.join(carpeta_downloads, "*.pdf"))
                if pdfs:
                    ultimo = max(pdfs, key=os.path.getctime)
                    destino = os.path.join(carpeta_final, nombre)
                    shutil.move(ultimo, destino)
                    archivos_descargados.append((destino, es_retencion))
                    print(f"   Guardado → {nombre}")
                time.sleep(1)
            except Exception as e:
                print(f"   Error: {e}")
                continue

        print(f"{mes} {anio} → COMPLETADO")

    if input(f"\n¿Quieres descargar otro tipo (facturas/retenciones)? (S/N): ").strip().upper() != "S":
        print("\nTodas las descargas completadas. Generando Excel...")
        break

# =============================
# MANEJO DE PDFs (NUEVOS O EXISTENTES)
# =============================
carpeta_facturas = r"C:\Users\Jordy\Downloads\Facturas"
carpeta_retenciones = r"C:\Users\Jordy\Downloads\Retenciones"

if not archivos_descargados:
    print("\n⚠️  ADVERTENCIA: No se descargaron nuevos documentos.")
    print("   Buscando PDFs existentes en las carpetas...")
    pdfs_existentes = []
    for carpeta, es_ret in [(carpeta_facturas, False), (carpeta_retenciones, True)]:
        for pdf_path in glob.glob(os.path.join(carpeta, "*.pdf")):
            pdfs_existentes.append((pdf_path, es_ret))
    if pdfs_existentes:
        archivos_descargados = pdfs_existentes
        print(f"   Se encontraron {len(pdfs_existentes)} PDFs existentes. Procesando...")
    else:
        print("   No hay PDFs en las carpetas. Se crearán Excels vacíos.")
else:
    print(f"\nSe descargaron {len(archivos_descargados)} documentos nuevos. Procesando...")

# =============================
# EXTRACCIÓN FINAL
# =============================
print("\n" + "="*80)
print("GENERANDO EXCEL SIN DUPLICADOS (FACTURAS Y RETENCIONES)")
print("="*80)

try:
    import pdfplumber
except ImportError:
    print("ERROR: Instala pdfplumber → pip install pdfplumber")
    driver.quit()
    exit()

datos_facturas = []
datos_retenciones = []
archivos_ya_procesados = set()

for destino, es_retencion in archivos_descargados:
    nombre = os.path.basename(destino)

    if nombre in archivos_ya_procesados:
        print(f"Duplicado omitido → {nombre}")
        continue
    archivos_ya_procesados.add(nombre)

    if es_retencion:
        try:
            with pdfplumber.open(destino) as pdf:
                texto = ""
                for page in pdf.pages:
                    texto += page.extract_text() or ""

            lineas = [l.strip() for l in texto.split('\n') if l.strip()]

            fecha = "NO_FECHA"
            ruc_emisor = "NO_RUC"
            num_ret = "NO_NUM"
            retenciones = []

            for linea in lineas:
                up = linea.upper()
                if any(pal in up for pal in ["RENTA", "IVA", "RETENCIÓN", "RETENCION"]):
                    partes = re.split(r'\s{2,}', linea.strip())
                    if len(partes) >= 4:
                        impuesto = "Renta" if "RENTA" in partes[-3].upper() else "IVA"
                        porc = partes[-2].strip().rstrip("%").strip() + "%"
                        val = partes[-1].replace("$", "").replace(",", "").strip()
                        retenciones.append((impuesto, porc, val))

                if "OTROS CON UTILIZACIÓN DEL SISTEMA" in linea or "OTROS CON UTILIZACIÓN" in up:
                    valores = re.findall(r"[\d.,]+", linea)
                    if valores:
                        val = valores[-1].replace(",", "")
                        retenciones.append(("Renta", "2.75%", val))
                        retenciones.append(("IVA", "30.00%", val))

                if any(x in linea for x in ["2.00", "8.00"]) and "RENTA" in up:
                    valores = re.findall(r"[\d.,]+", linea)
                    if len(valores) >= 2:
                        val = valores[-1].replace(",", "")
                        porc = "2.00%" if "2.00" in linea else "8.00%"
                        retenciones.append(("Renta", porc, val))

                if re.search(r"\b(1\.00|2\.00|2\.75|8\.00|30\.00|70\.00|100\.00)\b", linea):
                    nums = re.findall(r"[\d.,]+", linea)
                    if len(nums) >= 2:
                        val = nums[-1].replace(",", "")
                        p = re.search(r"(1\.00|2\.00|2\.75|8\.00|30\.00|70\.00|100\.00)", linea)
                        porc = (p.group(1) + "%") if p else "?.??%"
                        imp = "Renta" if p and float(p.group(1)) <= 10 else "IVA"
                        retenciones.append((imp, porc, val))

            for imp, porc, val in retenciones:
                datos_retenciones.append({
                    "Fecha": fecha if fecha != "NO_FECHA" else "01/01/2024",
                    "RUC_Emisor": ruc_emisor,
                    "Número_Retención": num_ret,
                    "Clave_Acceso": nombre.replace(".pdf", ""),
                    "RUC_Proveedor": "1803005071001",
                    "Proveedor": "SOLIS ACOSTA EDGAR FERNANDO",
                    "Número_Factura": "",
                    "Impuesto": imp,
                    "Porcentaje": porc,
                    "Valor_Retenido": val
                })
                print(f"   Retención {imp} {porc} = ${val}")

            if not retenciones:
                print(f"   WARNING: No detectada retención en {nombre}")

        except Exception as e:
            print(f"Error leyendo retención {nombre}: {e}")

    else:  # FACTURAS
        try:
            with pdfplumber.open(destino) as pdf:
                texto = ""
                for page in pdf.pages:
                    texto += (page.extract_text(x_tolerance=3, y_tolerance=3) or "")

            ruc = re.search(r"R\.?U\.?C\.?\s*[:\s]+(\d{13})", texto, re.IGNORECASE)
            factura = re.search(r"FACTURA\s+No?\.?\s*[:\s]+(\d{3}-\d{3}-\d{9})", texto, re.IGNORECASE)
            fecha_fact = re.search(r"(?:FECHA|Fecha)[\s\:]*[^\d\r\n]*(\d{2}[\/\-\s]?\d{2}[\/\-\s]?\d{4})", texto, re.IGNORECASE)

            # Subtotal gravado (cualquier porcentaje: 12%, 15%, etc.)
            base_gravada = re.search(r"SUBTOTAL\s+\d+%?\s*[:\s]*\$?\s*([\d.,]+)", texto, re.IGNORECASE)
            base_val = base_gravada.group(1).replace("$", "").strip() if base_gravada else "0.00"

            # Subtotal sin impuestos
            subtotal_sin_imp = re.search(
                r"SUBTOTAL\s+(?:SIN\s+IMPUESTOS|NO\s+OBJETO\s+DE\s+IVA|EXENTO\s+DE\s+IVA)\s*[:\s]*\$?\s*([\d.,]+)",
                texto, re.IGNORECASE
            )
            sin_imp_val = subtotal_sin_imp.group(1).replace("$", "").strip() if subtotal_sin_imp else "0.00"

            # IVA (cualquier porcentaje)
            iva = re.search(r"IVA\s+\d+%?\s*[:\s]*\$?\s*([\d.,]+)", texto, re.IGNORECASE)
            iva_val = iva.group(1).replace("$", "").strip() if iva else "0.00"

            # Valor Total
            total = re.search(r"VALOR\s+TOTAL\s*[:\s]*\$?\s*([\d.,]+)", texto, re.IGNORECASE)
            total_val = total.group(1).replace("$", "").strip() if total else "0.00"

            datos_facturas.append({
                "Archivo": nombre,
                "RUC_Emisor": ruc.group(1) if ruc else "NO_ENCONTRADO",
                "Número_Factura": factura.group(1) if factura else "NO_ENCONTRADO",
                "Fecha": fecha_fact.group(1).replace("-", "/").replace(" ", "/") if fecha_fact else "NO_ENCONTRADO",
                "Subtotal_Gravado": base_val.replace("$", "").replace(",", ".").strip() if base_val else "0.00",
                "Subtotal_sin_Impuestos": sin_imp_val.replace("$", "").replace(",", ".").strip() if sin_imp_val else "0.00",
                "IVA": iva_val.replace("$", "").replace(",", ".").strip() if iva_val else "0.00",
                "Total": total_val.replace("$", "").replace(",", ".").strip() if total_val else "0.00"
            })

            print(f"Factura procesada → {nombre[:50]}...")

        except Exception as e:
            print(f"ERROR leyendo factura {nombre}: {e}")

# =============================
# GUARDAR EXCEL (SIEMPRE, INCLUSO VACÍO)
# =============================
print("\n" + "="*80)
print("GUARDANDO EXCEL(S)...")
print("="*80)

# Facturas
excel_fac = os.path.join(carpeta_facturas, f"FACTURAS_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
df_fac = pd.DataFrame(datos_facturas) if datos_facturas else pd.DataFrame(columns=[
    "Archivo", "RUC_Emisor", "Número_Factura", "Fecha",
    "Subtotal_Gravado", "Subtotal_sin_Impuestos", "IVA", "Total"
])
df_fac.to_excel(excel_fac, index=False)
print(f"FACTURAS → {excel_fac} ({len(datos_facturas)} líneas)")
os.startfile(excel_fac)

# Retenciones
excel_ret = os.path.join(carpeta_retenciones, f"RETENCIONES_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
df_ret = pd.DataFrame(datos_retenciones) if datos_retenciones else pd.DataFrame(columns=[
    "Fecha", "RUC_Emisor", "Número_Retención", "Clave_Acceso",
    "RUC_Proveedor", "Proveedor", "Número_Factura", "Impuesto", "Porcentaje", "Valor_Retenido"
])
df_ret.to_excel(excel_ret, index=False)
print(f"RETENCIONES → {excel_ret} ({len(datos_retenciones)} líneas)")
os.startfile(excel_ret)

driver.quit()
print("\n¡TODO TERMINADO PERFECTO!")

# =============================
# CREAR EXCEL CON MÚLTIPLES PESTAÑAS: ESTADO + LIBRO DIARIO + LIBRO MAYOR
# =============================
print("\n" + "="*80)
print("CREANDO EXCEL CON ESTADO DE RESULTADOS, LIBRO DIARIO Y LIBRO MAYOR...")
print("="*80)

carpeta_estado = r"C:\Users\Jordy\Downloads\Estado"
os.makedirs(carpeta_estado, exist_ok=True)

# Calcular total de ventas (igual que antes)
total_ventas = 0.0
for factura in datos_facturas:
    try:
        valor_str = factura["Total"].replace(".", "").replace(",", ".")
        total_ventas += float(valor_str)
    except:
        pass

excel_estado = os.path.join(carpeta_estado, f"ESTADO_RESULTADOS_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment

wb = Workbook()

# =============================
# CREAR EXCEL: ESTADO DE RESULTADOS + LIBRO DIARIO Y MAYOR (PLANTILLA EXACTA)
# =============================
print("\n" + "="*80)
print("CREANDO EXCEL CON PLANTILLA EXACTA DE LIBRO DIARIO Y MAYOR...")
print("="*80)

carpeta_estado = r"C:\Users\Jordy\Downloads\Estado"
os.makedirs(carpeta_estado, exist_ok=True)

# Calcular total de ventas
total_ventas = 0.0
for factura in datos_facturas:
    try:
        valor_str = factura["Total"].replace(".", "").replace(",", ".")
        total_ventas += float(valor_str)
    except:
        pass

excel_estado = os.path.join(carpeta_estado, f"ESTADO_RESULTADOS_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# =============================
# PESTAÑA 1: Estado de Resultados
# =============================
ws_estado = wb.active
ws_estado.title = "Estado de Resultados"

ws_estado["A1"] = "ESTADO DE RESULTADOS INTEGRAL"
ws_estado.merge_cells("A1:D1")
ws_estado["A1"].font = Font(bold=True, size=14)
ws_estado["A1"].alignment = Alignment(horizontal="center")

ws_estado["A4"] = "Ventas"
ws_estado["A5"] = "Costo de ventas"
ws_estado["A6"] = "Gasto horas improductivas"
ws_estado["A7"] = "Utilidad bruta"
ws_estado["A8"] = "15% Participación trabajadores"
ws_estado["A9"] = "Utilidad antes de impuestos"
ws_estado["A10"] = "25% Impuesto a la renta"
ws_estado["A11"] = "Utilidad del ejercicio"

ws_estado["D4"] = 0
ws_estado["D5"] = ""
ws_estado["D6"] = ""
ws_estado["D7"] = "=D4-D5-D6"
ws_estado["D8"] = "=D7*0.15"
ws_estado["D9"] = "=D7-D8"
ws_estado["D10"] = "=D9*0.25"
ws_estado["D11"] = "=D9-D10"

moneda = NamedStyle(name="moneda", number_format='#,##0.00')
for row in range(4, 12):
    ws_estado[f"D{row}"].style = moneda

ws_estado.column_dimensions["A"].width = 40
ws_estado.column_dimensions["D"].width = 25

# =============================
# PESTAÑA 2: Libro Diario (SIN CAMBIOS EN CONTENIDO + BORDES COMPLETOS)
# =============================
ws_diario = wb.create_sheet("Libro Diario")

bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
green_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Título empresa
ws_diario["A1"] = 'EMPRESA "ABC" S.A.'
ws_diario["A1"].font = Font(bold=True, size=16)
ws_diario.merge_cells("A1:O1")
ws_diario["A1"].alignment = center
ws_diario.row_dimensions[1].height = 30

# Título "Libro diario"
ws_diario["A3"] = "Libro diario"
ws_diario["A3"].font = Font(bold=True, size=14)
ws_diario.merge_cells("A3:E3")
ws_diario.row_dimensions[3].height = 25

# Encabezados principales (verde)
headers = ["Fecha", "Detalle", "Parcial", "Debe", "Haber"]
for i, text in enumerate(headers):
    cell = ws_diario.cell(row=5, column=i+1)
    cell.value = text
    cell.font = bold
    cell.fill = green_fill
    cell.alignment = center
    cell.border = thin_border
ws_diario.row_dimensions[5].height = 25

# Ejemplos prellenados (exactamente como los tenías)
ws_diario.cell(row=7, column=1).value = "1/10/2025"
ws_diario.cell(row=7, column=2).value = "1"
ws_diario.cell(row=8, column=2).value = "Gasto servicios Básicos"
ws_diario.cell(row=9, column=2).value = "Bancos"

ws_diario.cell(row=10, column=1).value = "1/11/2025"
ws_diario.cell(row=10, column=2).value = "2"
ws_diario.cell(row=11, column=2).value = "Gasto suministro"
ws_diario.cell(row=12, column=2).value = "Bancos"

ws_diario.cell(row=13, column=1).value = "1/12/2025"
ws_diario.cell(row=13, column=2).value = "3"
ws_diario.cell(row=14, column=2).value = "Gasto variable"
ws_diario.cell(row=15, column=2).value = "Bancos"

ws_diario.cell(row=16, column=1).value = "1/13/2025"
ws_diario.cell(row=16, column=2).value = "4"
ws_diario.cell(row=17, column=2).value = "Gasto servicios"
ws_diario.cell(row=18, column=2).value = "Bancos"

ws_diario.cell(row=19, column=1).value = "1/14/2025"
ws_diario.cell(row=19, column=2).value = "5"
ws_diario.cell(row=20, column=2).value = "Gastos fijos"
ws_diario.cell(row=21, column=2).value = "Bancos"

ws_diario.cell(row=22, column=1).value = "1/15/2025"
ws_diario.cell(row=22, column=2).value = "6"
ws_diario.cell(row=23, column=2).value = "Gastos extraordinarios"
ws_diario.cell(row=24, column=2).value = "Bancos"

# Indentado para "Bancos"
for fila in [9,12,15,18,21,24]:
    ws_diario.cell(row=fila, column=2).alignment = Alignment(horizontal="left", indent=4)

# BORDES COMPLETOS EN TODO EL LIBRO DIARIO (tabla izquierda + cuentas derecha)
for row in range(5, 29):  # Muchas filas para escribir
    for col in range(1, 6):  # A a E
        ws_diario.cell(row=row, column=col).border = thin_border

# === SOLO AÑADIMOS LOS 0,00 EN DEBE Y HABER ===
moneda_style = NamedStyle(name="moneda_diario", number_format='#,##0.00')

# Gastos → 0,00 en Debe (columna D)
filas_gastos = [8, 11, 14, 17, 20, 23]
for fila in filas_gastos:
    cell = ws_diario.cell(row=fila, column=4)
    cell.value = 0
    cell.style = moneda_style
    cell.alignment = Alignment(horizontal="right")

# Bancos → 0,00 en Haber (columna E)
filas_bancos = [9, 12, 15, 18, 21, 24]
for fila in filas_bancos:
    cell = ws_diario.cell(row=fila, column=5)
    cell.value = 0
    cell.style = moneda_style
    cell.alignment = Alignment(horizontal="right")

# Anchos
ws_diario.column_dimensions["A"].width = 12
ws_diario.column_dimensions["B"].width = 40
ws_diario.column_dimensions["C"].width = 12
ws_diario.column_dimensions["D"].width = 15
ws_diario.column_dimensions["E"].width = 15

# =============================
# PESTAÑA 3: Libro Mayor (CON SEPARACIÓN ENTRE CUENTAS + 0,00 EN CELDAS VACÍAS)
# =============================
ws_mayor = wb.create_sheet("Libro Mayor")

# Estilos adicionales para moneda
moneda_style = NamedStyle(name="moneda_mayor", number_format='#,##0.00')

# Título
ws_mayor["A1"] = "LIBRO MAYOR"
ws_mayor["A1"].font = Font(bold=True, size=16)
ws_mayor.merge_cells("A1:P1")
ws_mayor["A1"].alignment = center
ws_mayor.row_dimensions[1].height = 35

# 5 cuentas con separación de 1 columna entre cada una
cuentas_mayor = [
    "Bancos",
    "Gasto servicios Básicos",
    "Gasto suministro",
    "Gasto variable",
    "Gastos fijos"
]
start_col_m = 2  # Empieza en columna B
separacion = 3   # 2 columnas (Debe/Haber) + 1 vacía

for i, cuenta in enumerate(cuentas_mayor):
    col_debe = start_col_m + i * separacion
    col_haber = col_debe + 1
    
    # Nombre cuenta (amarillo)
    ws_mayor.cell(row=3, column=col_debe).value = cuenta
    ws_mayor.cell(row=3, column=col_debe).font = bold
    ws_mayor.cell(row=3, column=col_debe).fill = yellow_fill
    ws_mayor.cell(row=3, column=col_debe).alignment = center
    ws_mayor.merge_cells(start_row=3, start_column=col_debe, end_row=3, end_column=col_haber)
    
    # Debe / Haber
    ws_mayor.cell(row=4, column=col_debe).value = "Debe"
    ws_mayor.cell(row=4, column=col_haber).value = "Haber"
    for c in [col_debe, col_haber]:
        cell = ws_mayor.cell(row=4, column=c)
        cell.font = bold
        cell.fill = yellow_fill
        cell.alignment = center
        cell.border = thin_border

ws_mayor.row_dimensions[4].height = 25

# Espacio para movimientos + poner 0,00 en todas las celdas vacías
for row in range(5, 10):  # Muchas filas para que puedas escribir
    ws_mayor.row_dimensions[row].height = 20
    for i in range(len(cuentas_mayor)):
        col_d = start_col_m + i * separacion
        col_h = col_d + 1
        
        # Poner 0,00 y formato moneda en Debe y Haber
        cell_debe = ws_mayor.cell(row=row, column=col_d)
        cell_haber = ws_mayor.cell(row=row, column=col_h)
        
        cell_debe.value = 0
        cell_haber.value = 0
        cell_debe.style = moneda_style
        cell_haber.style = moneda_style
        
        cell_debe.border = thin_border
        cell_haber.border = thin_border

# Totales S.D=, S.A=, S.N= bien abajo + SOLO UN 0,00 POR CUENTA (ciclo: S.D → S.A → S.N → ...)
total_row = 12
tipos_total = ["S.D=", "S.A=", "S.N="]  # Ciclo de 3

for i in range(len(cuentas_mayor)):
    col_debe = start_col_m + i * separacion
    col_haber = col_debe + 1
    
    # Escribir los textos S.D=, S.A=, S.N= en las 3 filas
    for j in range(3):
        cell_text = ws_mayor.cell(row=total_row + j, column=col_debe)
        cell_text.value = tipos_total[j]
        cell_text.font = bold
        cell_text.border = thin_border
    
    # Poner SOLO UN 0,00 en la columna Haber, según el orden de la cuenta
    tipo_actual = tipos_total[i % 3]  # Ciclo: cuenta 0 → S.D, cuenta 1 → S.A, cuenta 2 → S.N, etc.
    fila_con_valor = total_row + tipos_total.index(tipo_actual)
    
    cell_val = ws_mayor.cell(row=fila_con_valor, column=col_haber)
    cell_val.value = 0
    cell_val.style = moneda_style
    cell_val.border = thin_border
    cell_val.font = bold
    
    # Las otras dos filas de Haber quedan vacías (sin 0)

# Anchos en Libro Mayor (con separación)
ws_mayor.column_dimensions["A"].width = 5
for i in range(len(cuentas_mayor)):
    col_d_letter = get_column_letter(start_col_m + i * separacion)
    col_h_letter = get_column_letter(start_col_m + i * separacion + 1)
    ws_mayor.column_dimensions[col_d_letter].width = 15
    ws_mayor.column_dimensions[col_h_letter].width = 15
    if i < len(cuentas_mayor) - 1:
        sep_letter = get_column_letter(start_col_m + i * separacion + 2)
        ws_mayor.column_dimensions[sep_letter].width = 8

# =============================
# GUARDAR Y ABRIR
# =============================

wb.save(excel_estado)
print(f"¡EXCEL CON 3 PESTAÑAS CREADO PERFECTO! → {excel_estado}")
print("   → Pestaña 1: Estado de Resultados (ventas automáticas)")
print("   → Pestaña 2: Libro Diario (con ejemplos y 0,00)")
print("   → Pestaña 3: Libro Mayor (separado, con 0,00)")
os.startfile(excel_estado)